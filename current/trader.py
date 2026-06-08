from dataclasses import dataclass
from datetime import datetime, date
import csv
from collections import defaultdict


@dataclass
class Position:
    side: str
    entry_price: float
    entry_time: str
    reason: str
    contracts: int


class SimulatedTrader:
    def __init__(self, initial_equity: float, point_value: int, margin_per_contract: int, contracts: int):
        self.initial_equity = initial_equity
        self.realized_pnl = 0.0
        self.point_value = point_value
        self.margin_per_contract = margin_per_contract
        self.contracts = contracts
        self.position: Position | None = None
        self.trades = []
        self.open_count = 0

        self.trade_date = date.today()
        self.day_start_equity = float(initial_equity)
        self.day_start_realized_pnl = 0.0
        self.daily_risk_locked = False
        self.same_direction_notice_shown = False
        self.last_open_ts = 0.0
        self.last_close_ts = 0.0
        self.last_stop_loss_ts = 0.0

    def reset_daily_baseline_if_needed(self, current_price: float):
        today = date.today()
        if today != self.trade_date:
            self.trade_date = today
            self.day_start_equity = self.equity(current_price)
            self.day_start_realized_pnl = self.realized_pnl
            self.open_count = 0
            self.daily_risk_locked = False
            self.same_direction_notice_shown = False
            self.last_open_ts = 0.0
            self.last_close_ts = 0.0
            self.last_stop_loss_ts = 0.0

    def set_contract(self, point_value: int, margin_per_contract: int, contracts: int):
        self.point_value = point_value
        self.margin_per_contract = margin_per_contract
        self.contracts = contracts

    def floating_pnl(self, current_price: float) -> float:
        if self.position is None:
            return 0.0
        if self.position.side == "LONG":
            pnl_points = current_price - self.position.entry_price
        else:
            pnl_points = self.position.entry_price - current_price
        return pnl_points * self.point_value * self.position.contracts

    def equity(self, current_price: float) -> float:
        return self.initial_equity + self.realized_pnl + self.floating_pnl(current_price)

    def daily_pnl(self, current_price: float) -> float:
        return self.equity(current_price) - self.day_start_equity

    def daily_pnl_percent(self, current_price: float) -> float:
        if self.day_start_equity == 0:
            return 0.0
        return self.daily_pnl(current_price) / self.day_start_equity * 100

    def max_daily_loss_amount(self, settings: dict) -> float:
        unit = settings.get("max_loss_unit", "元")
        value = abs(float(settings.get("max_loss", 0)))
        if unit == "%":
            return self.day_start_equity * value / 100
        return value

    def used_margin(self) -> float:
        if self.position is None:
            return 0.0
        return self.margin_per_contract * self.position.contracts

    def available_funds(self, current_price: float) -> float:
        return self.equity(current_price) - self.used_margin()

    def risk_ratio(self, current_price: float) -> float:
        eq = self.equity(current_price)
        if eq <= 0:
            return 999.0
        return self.used_margin() / eq * 100

    def required_margin_for_new_position(self) -> float:
        return self.margin_per_contract * self.contracts

    def is_daily_loss_reached(self, settings: dict, current_price: float) -> bool:
        self.reset_daily_baseline_if_needed(current_price)
        limit_amount = self.max_daily_loss_amount(settings)
        if limit_amount <= 0:
            return False
        return self.daily_pnl(current_price) <= -limit_amount

    def is_max_trades_reached(self, settings: dict) -> bool:
        limit = int(settings.get("max_trades", 0))
        if limit <= 0:
            return False
        return self.open_count >= limit

    def _check_open_allowed(self, settings: dict, current_price: float) -> tuple[bool, str]:
        self.reset_daily_baseline_if_needed(current_price)
        now_ts = datetime.now().timestamp()

        if self.daily_risk_locked or self.is_daily_loss_reached(settings, current_price):
            self.daily_risk_locked = True
            return False, (
                f"已達每日最大虧損限制：日損益 {self.daily_pnl(current_price):,.0f} 元，"
                f"限制 {self.max_daily_loss_amount(settings):,.0f} 元。"
            )

        if settings.get("enable_time_filter", False):
            now_minutes = datetime.now().hour * 60 + datetime.now().minute
            trade_start = settings.get("trade_start_hour", 8) * 60 + settings.get("trade_start_minute", 45)
            stop_open = settings.get("stop_open_hour", 13) * 60 + settings.get("stop_open_minute", 30)
            if now_minutes < trade_start:
                return False, "尚未到允許開倉時間"
            if now_minutes >= stop_open:
                return False, "已超過停止開倉時間"

        if self.is_max_trades_reached(settings):
            return False, "已達每日最大交易次數"

        open_cd = float(settings.get("open_cooldown_seconds", 0))
        if open_cd > 0 and self.last_open_ts > 0 and now_ts - self.last_open_ts < open_cd:
            remain = int(open_cd - (now_ts - self.last_open_ts))
            return False, f"開倉冷卻中，剩餘約 {max(remain, 1)} 秒"

        stop_loss_cd = float(settings.get("stop_loss_reentry_cooldown_seconds", 0))
        if stop_loss_cd > 0 and self.last_stop_loss_ts > 0 and now_ts - self.last_stop_loss_ts < stop_loss_cd:
            remain = int(stop_loss_cd - (now_ts - self.last_stop_loss_ts))
            return False, f"停損後重新進場冷卻中，剩餘約 {max(remain, 1)} 秒"

        required = self.required_margin_for_new_position()
        available = self.available_funds(current_price)
        if available < required:
            return False, f"資金不足，需要保證金 {required:,.0f}，可用資金 {available:,.0f}"
        return True, ""

    def check_daily_risk(self, price: float, settings: dict):
        """達每日最大虧損時鎖住交易；若有持倉，先平倉。"""
        if self.daily_risk_locked and self.position is None:
            return None
        if self.is_daily_loss_reached(settings, price):
            self.daily_risk_locked = True
            if self.position is not None:
                pnl = self.close_position(price, "達到每日最大虧損，自動平倉")
                return f"達到每日最大虧損，已自動平倉 @ {price}，本筆損益 {pnl:.0f} 元"
            return "已達每日最大虧損限制，今日停止自動交易"
        return None

    def on_signal(self, action: str, price: float, reason: str, settings: dict):
        # 有持倉時，反向訊號可直接平倉；不再做平倉冷卻，避免妨礙風控。
        if self.position is not None:
            if self.position.side == "LONG" and action == "SELL":
                pnl = self.close_position(price, f"反向訊號平多：{reason}")
                return f"平多單 @ {price}，損益 {pnl:.0f} 元"
            if self.position.side == "SHORT" and action == "BUY":
                pnl = self.close_position(price, f"反向訊號平空：{reason}")
                return f"平空單 @ {price}，損益 {pnl:.0f} 元"
            if not self.same_direction_notice_shown:
                self.same_direction_notice_shown = True
                side_text = "多單" if self.position.side == "LONG" else "空單"
                return f"已有{side_text}持倉，後續同方向訊號將忽略"
            return None

        ok, msg = self._check_open_allowed(settings, price)
        if not ok:
            return msg

        if action == "BUY":
            self.open_position("LONG", price, reason)
            return f"開多單 {self.contracts}口 @ {price}，原因：{reason}"
        if action == "SELL":
            self.open_position("SHORT", price, reason)
            return f"開空單 {self.contracts}口 @ {price}，原因：{reason}"
        return None

    def check_stop(self, price: float, settings: dict):
        if self.position is None:
            return None
        if self.position.side == "LONG":
            diff = price - self.position.entry_price
        else:
            diff = self.position.entry_price - price
        if diff >= settings["take_profit"]:
            pnl = self.close_position(price, "達到停利")
            return f"停利平倉 @ {price}，損益 {pnl:.0f} 元"
        if diff <= -settings["stop_loss"]:
            pnl = self.close_position(price, "達到停損")
            self.last_stop_loss_ts = datetime.now().timestamp()
            return f"停損平倉 @ {price}，損益 {pnl:.0f} 元"
        return None

    def open_position(self, side: str, price: float, reason: str):
        self.same_direction_notice_shown = False
        self.last_open_ts = datetime.now().timestamp()
        self.position = Position(
            side=side,
            entry_price=price,
            entry_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            reason=reason,
            contracts=self.contracts,
        )
        self.open_count += 1

    def close_position(self, price: float, reason: str):
        if self.position is None:
            return 0.0
        if self.position.side == "LONG":
            pnl_points = price - self.position.entry_price
        else:
            pnl_points = self.position.entry_price - price
        pnl_money = pnl_points * self.point_value * self.position.contracts
        self.realized_pnl += pnl_money
        record = {
            "entry_time": self.position.entry_time,
            "exit_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "side": self.position.side,
            "contracts": self.position.contracts,
            "entry_price": self.position.entry_price,
            "exit_price": price,
            "pnl_points": pnl_points,
            "pnl_money": pnl_money,
            "equity_after": self.initial_equity + self.realized_pnl,
            "daily_pnl_after": self.realized_pnl - self.day_start_realized_pnl,
            "entry_reason": self.position.reason,
            "exit_reason": reason,
        }
        self.trades.append(record)
        self.position = None
        self.same_direction_notice_shown = False
        self.last_close_ts = datetime.now().timestamp()
        return pnl_money

    def check_force_close_time(self, price: float, settings: dict):
        if not settings.get("enable_time_filter", False):
            return None
        if not settings.get("enable_force_close", False):
            return None
        if self.position is None:
            return None
        now_minutes = datetime.now().hour * 60 + datetime.now().minute
        force_close = settings.get("force_close_hour", 13) * 60 + settings.get("force_close_minute", 44)
        if now_minutes >= force_close:
            pnl = self.close_position(price, "達到強制平倉時間")
            return f"達到強制平倉時間，已平倉 @ {price}，損益 {pnl:.0f} 元"
        return None

    def export_trades(self, filename="trades.csv"):
        if not self.trades:
            return False
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=self.trades[0].keys())
            writer.writeheader()
            writer.writerows(self.trades)
        return True

    def export_trades_excel(self, filename="trades.xlsx"):
        if not self.trades:
            return False
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl，請先執行 pip3 install --break-system-packages openpyxl") from exc

        wb = Workbook()

        # 工作表1：交易紀錄
        ws = wb.active
        ws.title = "交易紀錄"
        headers = [
            "進場時間", "出場時間", "方向", "口數", "進場價", "出場價",
            "損益點數", "損益金額", "出場後權益", "今日已實現損益", "進場原因", "出場原因"
        ]
        ws.append(headers)
        for trade in self.trades:
            ws.append([
                trade.get("entry_time", ""),
                trade.get("exit_time", ""),
                "多單" if trade.get("side") == "LONG" else "空單",
                trade.get("contracts", 0),
                trade.get("entry_price", 0),
                trade.get("exit_price", 0),
                trade.get("pnl_points", 0),
                trade.get("pnl_money", 0),
                trade.get("equity_after", 0),
                trade.get("daily_pnl_after", 0),
                trade.get("entry_reason", ""),
                trade.get("exit_reason", ""),
            ])

        # 工作表2：統計報表
        stats = self.stats()
        gross_profit = sum(t["pnl_money"] for t in self.trades if t["pnl_money"] > 0)
        gross_loss = sum(t["pnl_money"] for t in self.trades if t["pnl_money"] < 0)
        net_profit = sum(t["pnl_money"] for t in self.trades)
        ws2 = wb.create_sheet("統計報表")
        ws2.append(["項目", "數值"])
        summary_rows = [
            ("總完成交易", stats["closed_trades"]),
            ("今日開倉次數", stats["open_count"]),
            ("勝率", f"{stats['win_rate']:.2f}%"),
            ("平均獲利", round(stats["avg_win"], 2)),
            ("平均虧損", round(stats["avg_loss"], 2)),
            ("獲利因子", round(stats["profit_factor"], 4)),
            ("總獲利", round(gross_profit, 2)),
            ("總虧損", round(gross_loss, 2)),
            ("淨利潤", round(net_profit, 2)),
        ]
        for row in summary_rows:
            ws2.append(row)

        # 工作表3：每日統計
        daily = defaultdict(lambda: {"count": 0, "pnl": 0.0})
        for trade in self.trades:
            day = str(trade.get("exit_time", ""))[:10]
            daily[day]["count"] += 1
            daily[day]["pnl"] += float(trade.get("pnl_money", 0))
        ws3 = wb.create_sheet("每日統計")
        ws3.append(["日期", "完成交易次數", "損益金額"])
        for day in sorted(daily):
            ws3.append([day, daily[day]["count"], round(daily[day]["pnl"], 2)])

        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column_cells in sheet.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                width = min(max(max_len + 2, 10), 35)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

        wb.save(filename)
        return True

    def stats(self):
        total = len(self.trades)
        wins = [t for t in self.trades if t["pnl_money"] > 0]
        losses = [t for t in self.trades if t["pnl_money"] < 0]
        gross_profit = sum(t["pnl_money"] for t in wins)
        gross_loss = abs(sum(t["pnl_money"] for t in losses))
        return {
            "closed_trades": total,
            "open_count": self.open_count,
            "win_rate": (len(wins) / total * 100) if total else 0.0,
            "avg_win": (gross_profit / len(wins)) if wins else 0.0,
            "avg_loss": (gross_loss / len(losses)) if losses else 0.0,
            "profit_factor": (gross_profit / gross_loss) if gross_loss else 0.0,
        }

    def position_text(self):
        if self.position is None:
            return "無持倉"
        side = "多單" if self.position.side == "LONG" else "空單"
        return f"{side} {self.position.contracts}口 @ {self.position.entry_price}"
